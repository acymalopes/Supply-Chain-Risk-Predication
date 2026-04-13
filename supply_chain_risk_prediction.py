"""
================================================================
  Supply Chain Risk Prediction System
  Data Science Internship Project
================================================================
  Objective : Predict potential disruptions & risk levels
              within a supply chain using ML models.
  Models     : Random Forest | XGBoost | Logistic Regression
  Techniques : EDA, Feature Engineering, SMOTE, Model Comparison,
               Feature Importance, Risk Scoring
  Tools      : Python, NumPy, Pandas, Scikit-learn, XGBoost
================================================================
"""

import numpy as np
import pandas as pd
import warnings
warnings.filterwarnings("ignore")

from sklearn.model_selection   import train_test_split, cross_val_score, StratifiedKFold
from sklearn.preprocessing     import LabelEncoder, StandardScaler
from sklearn.ensemble          import RandomForestClassifier, GradientBoostingClassifier
from sklearn.linear_model      import LogisticRegression
from sklearn.metrics           import (classification_report, confusion_matrix,
                                       accuracy_score, roc_auc_score, f1_score)
from sklearn.pipeline          import Pipeline
from imblearn.over_sampling    import SMOTE
from xgboost                   import XGBClassifier
from openpyxl                  import Workbook
from openpyxl.styles           import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils            import get_column_letter


# ─────────────────────────────────────────────────────────────
# 1. LOAD & EXPLORE DATA
# ─────────────────────────────────────────────────────────────

FEATURE_COLS = [
    "Geopolitical_Score", "Weather_Severity", "Financial_Health",
    "Logistics_Delay_Days", "Quality_Reject_Rate", "Inventory_Days",
    "Demand_Volatility", "Port_Congestion", "Exchange_Rate_Risk",
    "Past_Disruptions", "Supplier_Dependency", "Compliance_Score"
]

def load_and_explore(filepath="supply_chain_data.xlsx"):
    xl   = pd.read_excel(filepath, sheet_name=None)
    df   = xl["Supply_Chain_Events"]
    sups = xl["Suppliers"]
    df.columns   = df.columns.str.replace(" ", "_")
    sups.columns = sups.columns.str.replace(" ", "_")

    df["Date"]     = pd.to_datetime(df["Date"], errors="coerce")
    df["Month"]    = df["Date"].dt.month
    df["Quarter"]  = df["Date"].dt.quarter
    df["Year"]     = df["Date"].dt.year

    print("=" * 60)
    print("  SUPPLY CHAIN RISK PREDICTION — DATA OVERVIEW")
    print("=" * 60)
    print(f"  Total Records   : {len(df)}")
    print(f"  Suppliers       : {df['Supplier_ID'].nunique()}")
    print(f"  Countries       : {df['Country'].nunique()}")
    print(f"  Date Range      : {df['Date'].min().date()} → {df['Date'].max().date()}")
    print(f"  Disruption Rate : {df['Disruption_Occurred'].mean()*100:.1f}%")
    print()
    print("  Risk Level Distribution:")
    print(df["Risk_Level"].value_counts().to_string())
    print()
    print("  Top 5 High-Risk Countries:")
    top_c = (df[df["Risk_Level"]=="High"]
             .groupby("Country")
             .size()
             .sort_values(ascending=False)
             .head(5))
    print(top_c.to_string())
    print()

    return df, sups


# ─────────────────────────────────────────────────────────────
# 2. FEATURE ENGINEERING
# ─────────────────────────────────────────────────────────────

def engineer_features(df):
    """Create additional derived risk features."""

    # Supply Chain Vulnerability Index
    df["Vulnerability_Index"] = (
        df["Supplier_Dependency"] / 100 * 0.3 +
        (10 - df["Financial_Health"]) / 10 * 0.3 +
        df["Past_Disruptions"] / 8 * 0.2 +
        df["Logistics_Delay_Days"] / 60 * 0.2
    )

    # Operational Stress Score
    df["Operational_Stress"] = (
        df["Quality_Reject_Rate"] / 15 * 0.4 +
        (30 / (df["Inventory_Days"] + 1)) * 0.3 +
        df["Demand_Volatility"] / 10 * 0.3
    )

    # External Risk Score
    df["External_Risk"] = (
        df["Geopolitical_Score"] / 10 * 0.35 +
        df["Weather_Severity"] / 10 * 0.25 +
        df["Port_Congestion"] / 10 * 0.25 +
        df["Exchange_Rate_Risk"] / 10 * 0.15
    )

    # Low inventory flag
    df["Low_Inventory_Flag"] = (df["Inventory_Days"] < 15).astype(int)

    # High dependency flag
    df["High_Dependency_Flag"] = (df["Supplier_Dependency"] > 70).astype(int)

    # Country risk encoding
    high_risk_countries  = ["China", "Vietnam", "Mexico", "India"]
    low_risk_countries   = ["Germany", "Japan", "USA", "South Korea", "Taiwan"]
    df["Country_Risk"] = df["Country"].apply(
        lambda x: 2 if x in high_risk_countries else (0 if x in low_risk_countries else 1)
    )

    # Tier encoding
    tier_map = {"Tier 1": 1, "Tier 2": 2, "Tier 3": 3}
    df["Tier_Numeric"] = df["Risk_Tier"].map(tier_map)

    print("[✓] Feature Engineering complete — added 7 new features")
    return df


# ─────────────────────────────────────────────────────────────
# 3. PREPARE DATA FOR MODELLING
# ─────────────────────────────────────────────────────────────

EXTENDED_FEATURES = FEATURE_COLS + [
    "Vulnerability_Index", "Operational_Stress", "External_Risk",
    "Low_Inventory_Flag", "High_Dependency_Flag", "Country_Risk",
    "Tier_Numeric", "Month", "Quarter"
]

def prepare_data(df):
    X = df[EXTENDED_FEATURES].copy()
    y_binary     = df["Disruption_Occurred"]          # Binary: 0/1
    y_multiclass = df["Risk_Level"]                   # Multiclass: Low/Medium/High

    le = LabelEncoder()
    y_multi_enc = le.fit_transform(y_multiclass)

    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)
    X_scaled = pd.DataFrame(X_scaled, columns=EXTENDED_FEATURES)

    # Handle class imbalance with SMOTE
    smote = SMOTE(random_state=42)
    X_res, y_res = smote.fit_resample(X_scaled, y_binary)

    X_train, X_test, y_train, y_test = train_test_split(
        X_res, y_res, test_size=0.2, random_state=42, stratify=y_res
    )

    # Multiclass split (original, no SMOTE)
    X_sc_orig = pd.DataFrame(scaler.transform(df[EXTENDED_FEATURES]), columns=EXTENDED_FEATURES)
    Xm_train, Xm_test, ym_train, ym_test = train_test_split(
        X_sc_orig, y_multi_enc, test_size=0.2, random_state=42, stratify=y_multi_enc
    )

    print(f"[✓] Data Preparation complete")
    print(f"    Train size (binary)     : {len(X_train)}")
    print(f"    Test size  (binary)     : {len(X_test)}")
    print(f"    Features used           : {len(EXTENDED_FEATURES)}")
    print(f"    Class balance after SMOTE: {dict(pd.Series(y_res).value_counts())}")
    print()

    return (X_train, X_test, y_train, y_test,
            Xm_train, Xm_test, ym_train, ym_test,
            scaler, le, X_scaled, df[EXTENDED_FEATURES])


# ─────────────────────────────────────────────────────────────
# 4. TRAIN & EVALUATE MODELS
# ─────────────────────────────────────────────────────────────

def train_models(X_train, X_test, y_train, y_test):
    print("=" * 60)
    print("  MODEL TRAINING & EVALUATION (Binary: Disruption)")
    print("=" * 60)

    models = {
        "Logistic Regression": LogisticRegression(max_iter=1000, random_state=42, C=0.5),
        "Random Forest":       RandomForestClassifier(n_estimators=200, max_depth=8,
                                                      random_state=42, n_jobs=-1),
        "XGBoost":             XGBClassifier(n_estimators=200, max_depth=6, learning_rate=0.05,
                                             use_label_encoder=False, eval_metric="logloss",
                                             random_state=42, verbosity=0),
        "Gradient Boosting":   GradientBoostingClassifier(n_estimators=150, max_depth=5,
                                                          learning_rate=0.05, random_state=42),
    }

    results = {}
    cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)

    for name, model in models.items():
        model.fit(X_train, y_train)
        y_pred  = model.predict(X_test)
        y_proba = model.predict_proba(X_test)[:, 1] if hasattr(model, "predict_proba") else None

        acc    = accuracy_score(y_test, y_pred)
        f1     = f1_score(y_test, y_pred, average="weighted")
        auc    = roc_auc_score(y_test, y_proba) if y_proba is not None else None
        cv_sc  = cross_val_score(model, X_train, y_train, cv=cv, scoring="f1_weighted")

        results[name] = {
            "model":    model,
            "accuracy": round(acc, 4),
            "f1_score": round(f1, 4),
            "roc_auc":  round(auc, 4) if auc else None,
            "cv_mean":  round(cv_sc.mean(), 4),
            "cv_std":   round(cv_sc.std(), 4),
            "y_pred":   y_pred,
            "y_proba":  y_proba,
        }

        print(f"\n  [{name}]")
        print(f"    Accuracy  : {acc*100:.2f}%")
        print(f"    F1 Score  : {f1:.4f}")
        print(f"    ROC-AUC   : {auc:.4f}" if auc else "    ROC-AUC   : N/A")
        print(f"    CV F1     : {cv_sc.mean():.4f} ± {cv_sc.std():.4f}")
        print(f"    Confusion Matrix:\n{confusion_matrix(y_test, y_pred)}")

    # Best model
    best_name = max(results, key=lambda k: results[k]["f1_score"])
    print(f"\n  ★ Best Model: {best_name} (F1={results[best_name]['f1_score']})")
    return results, best_name


def train_multiclass(Xm_train, Xm_test, ym_train, ym_test, le):
    print("\n" + "=" * 60)
    print("  MULTICLASS MODEL (Risk Level: Low / Medium / High)")
    print("=" * 60)

    xgb_multi = XGBClassifier(n_estimators=200, max_depth=6, learning_rate=0.05,
                               objective="multi:softprob", num_class=3,
                               use_label_encoder=False, eval_metric="mlogloss",
                               random_state=42, verbosity=0)
    xgb_multi.fit(Xm_train, ym_train)
    ym_pred = xgb_multi.predict(Xm_test)
    acc = accuracy_score(ym_test, ym_pred)
    print(f"  XGBoost Multiclass Accuracy: {acc*100:.2f}%")
    print(f"\n{classification_report(ym_test, ym_pred, target_names=le.classes_)}")
    return xgb_multi


# ─────────────────────────────────────────────────────────────
# 5. FEATURE IMPORTANCE
# ─────────────────────────────────────────────────────────────

def feature_importance(results, best_name):
    model = results[best_name]["model"]
    if hasattr(model, "feature_importances_"):
        imp = pd.Series(model.feature_importances_, index=EXTENDED_FEATURES)
        imp = imp.sort_values(ascending=False)
        print("\n  Top 10 Feature Importances:")
        print(imp.head(10).to_string())
        return imp
    return None


# ─────────────────────────────────────────────────────────────
# 6. RISK SCORING — PREDICT ON ALL SUPPLIERS
# ─────────────────────────────────────────────────────────────

def score_all_suppliers(df, results, best_name, scaler):
    model = results[best_name]["model"]
    X_all = pd.DataFrame(scaler.transform(df[EXTENDED_FEATURES]), columns=EXTENDED_FEATURES)
    proba = model.predict_proba(X_all)[:, 1]
    pred  = model.predict(X_all)

    df = df.copy()
    df["Predicted_Disruption"]    = pred
    df["Disruption_Probability"]  = np.round(proba, 4)

    df["Predicted_Risk_Category"] = pd.cut(
        df["Disruption_Probability"],
        bins=[-0.001, 0.30, 0.60, 1.001],
        labels=["Low Risk", "Medium Risk", "High Risk"]
    )

    print(f"\n[✓] Scored {len(df)} records")
    print("    Predicted Risk Distribution:")
    print(df["Predicted_Risk_Category"].value_counts().to_string())
    return df


# ─────────────────────────────────────────────────────────────
# 7. EXPORT RESULTS TO EXCEL
# ─────────────────────────────────────────────────────────────

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

def _hdr(ws, cols, bg="0D1B2A"):
    fill = PatternFill("solid", start_color=bg)
    ws.row_dimensions[1].height = 28
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=col.replace("_", " "))
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN

def _row(ws, ri, values, bg="FFFFFF"):
    fill = PatternFill("solid", start_color=bg)
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.fill = fill
        c.border = THIN
        c.alignment = Alignment(horizontal="center", vertical="center")

def _col_widths(ws, df, cols):
    for ci, col in enumerate(cols, 1):
        mx = max(len(str(col)), df[col].astype(str).str.len().max() if col in df.columns else 10)
        ws.column_dimensions[get_column_letter(ci)].width = min(mx + 4, 30)

def export_results(scored_df, results, best_name, importance,
                   out="supply_chain_risk_results.xlsx"):
    wb = Workbook()

    RISK_COLORS = {
        "High Risk":   "FF6B6B",
        "Medium Risk": "FFD93D",
        "Low Risk":    "C8F7C5",
    }

    # ── Sheet 1: All Predictions ──────────────────────────────
    ws1 = wb.active
    ws1.title = "All_Predictions"
    out_cols = [
        "Record_ID", "Supplier_Name", "Country", "Category", "Risk_Tier",
        "Date", "Composite_Risk_Score", "Risk_Level",
        "Disruption_Occurred", "Disruption_Probability",
        "Predicted_Disruption", "Predicted_Risk_Category"
    ]
    out_cols = [c for c in out_cols if c in scored_df.columns]
    _hdr(ws1, out_cols, "0D1B2A")
    for ri, (_, row) in enumerate(scored_df[out_cols].iterrows(), 2):
        risk_cat = str(row.get("Predicted_Risk_Category", ""))
        bg = RISK_COLORS.get(risk_cat, "FFFFFF")
        alt = "EAF2FF" if ri % 2 == 0 else bg
        _row(ws1, ri, [row[c] for c in out_cols], bg=bg if risk_cat else alt)
    _col_widths(ws1, scored_df, out_cols)

    # ── Sheet 2: High Risk Records ────────────────────────────
    ws2 = wb.create_sheet("High_Risk_Records")
    hr  = scored_df[scored_df["Predicted_Risk_Category"] == "High Risk"]
    _hdr(ws2, out_cols, "922B21")
    for ri, (_, row) in enumerate(hr[out_cols].iterrows(), 2):
        _row(ws2, ri, [row[c] for c in out_cols], bg="FADBD8" if ri%2==0 else "FDECEA")
    _col_widths(ws2, hr, out_cols)

    # ── Sheet 3: Model Comparison ─────────────────────────────
    ws3 = wb.create_sheet("Model_Comparison")
    mc_cols = ["Model", "Accuracy", "F1_Score", "ROC_AUC", "CV_Mean", "CV_Std", "Best"]
    _hdr(ws3, mc_cols, "1A5276")
    for ri, (name, res) in enumerate(results.items(), 2):
        is_best = "★ BEST" if name == best_name else ""
        bg = "D5F5E3" if name == best_name else ("EAF2FF" if ri%2==0 else "FFFFFF")
        _row(ws3, ri, [
            name,
            f"{res['accuracy']*100:.2f}%",
            f"{res['f1_score']:.4f}",
            f"{res['roc_auc']:.4f}" if res['roc_auc'] else "N/A",
            f"{res['cv_mean']:.4f}",
            f"{res['cv_std']:.4f}",
            is_best
        ], bg=bg)
    for ci in range(1, 8):
        ws3.column_dimensions[get_column_letter(ci)].width = 22

    # ── Sheet 4: Feature Importance ───────────────────────────
    ws4 = wb.create_sheet("Feature_Importance")
    fi_cols = ["Rank", "Feature", "Importance_Score", "Importance_Pct"]
    _hdr(ws4, fi_cols, "4A235A")
    if importance is not None:
        total = importance.sum()
        for ri, (feat, imp) in enumerate(importance.items(), 2):
            bg = "EAF2FF" if ri%2==0 else "FFFFFF"
            _row(ws4, ri, [
                ri-1, feat.replace("_"," "),
                round(imp, 6), f"{imp/total*100:.2f}%"
            ], bg=bg)
    for ci in range(1, 5):
        ws4.column_dimensions[get_column_letter(ci)].width = 28

    # ── Sheet 5: Supplier Risk Summary ────────────────────────
    ws5 = wb.create_sheet("Supplier_Risk_Summary")
    sup_summary = (scored_df.groupby("Supplier_Name").agg(
        Avg_Risk_Score   = ("Composite_Risk_Score", "mean"),
        Avg_Disruption_P = ("Disruption_Probability", "mean"),
        Total_Events     = ("Record_ID", "count"),
        Disruptions      = ("Disruption_Occurred", "sum"),
        Disruption_Rate  = ("Disruption_Occurred", "mean"),
        Country          = ("Country", "first"),
        Category         = ("Category", "first"),
    ).reset_index().sort_values("Avg_Disruption_P", ascending=False).round(3))
    sup_cols = list(sup_summary.columns)
    _hdr(ws5, sup_cols, "145A32")
    for ri, (_, row) in enumerate(sup_summary.iterrows(), 2):
        prob = row["Avg_Disruption_P"]
        bg = "FF6B6B" if prob > 0.6 else ("FFD93D" if prob > 0.3 else "C8F7C5")
        _row(ws5, ri, [row[c] for c in sup_cols], bg=bg)
    _col_widths(ws5, sup_summary, sup_cols)

    # ── Sheet 6: Executive Dashboard ──────────────────────────
    ws6 = wb.create_sheet("Executive_Dashboard")
    ws6.column_dimensions["A"].width = 35
    ws6.column_dimensions["B"].width = 22

    title_cell = ws6.cell(row=1, column=1,
                          value="Supply Chain Risk Prediction — Executive Summary")
    title_cell.font = Font(bold=True, size=14, color="0D1B2A", name="Arial")
    ws6.merge_cells("A1:B1")

    hr_count = len(scored_df[scored_df["Predicted_Risk_Category"]=="High Risk"])
    mr_count = len(scored_df[scored_df["Predicted_Risk_Category"]=="Medium Risk"])
    lr_count = len(scored_df[scored_df["Predicted_Risk_Category"]=="Low Risk"])

    kpis = [
        ("Total Supply Chain Records",         len(scored_df)),
        ("Unique Suppliers Analysed",          scored_df["Supplier_ID"].nunique()),
        ("Countries Covered",                  scored_df["Country"].nunique()),
        ("High Risk Records",                  hr_count),
        ("Medium Risk Records",                mr_count),
        ("Low Risk Records",                   lr_count),
        ("Overall Disruption Rate (%)",        f"{scored_df['Disruption_Occurred'].mean()*100:.1f}%"),
        ("Avg Disruption Probability (%)",     f"{scored_df['Disruption_Probability'].mean()*100:.1f}%"),
        ("Best Prediction Model",              best_name),
        ("Model F1 Score",                     f"{results[best_name]['f1_score']:.4f}"),
        ("Model ROC-AUC",                      f"{results[best_name]['roc_auc']:.4f}"),
        ("Top Risk Country", scored_df[scored_df["Predicted_Risk_Category"]=="High Risk"]
                              ["Country"].value_counts().idxmax()),
        ("Most Vulnerable Category", scored_df[scored_df["Predicted_Risk_Category"]=="High Risk"]
                              ["Category"].value_counts().idxmax()),
    ]

    fills = ["FADBD8","FDECEA","D5EEF9","D5F5E3","EAF2FF","FEF9E7",
             "E8DAEF","FDFEFE","D6EAF8","E9F7EF","FEF9E7","FAD7A0","F9EBEA"]
    for ri, ((label, value), bg) in enumerate(zip(kpis, fills), 3):
        lc = ws6.cell(row=ri, column=1, value=label)
        vc = ws6.cell(row=ri, column=2, value=value)
        lc.font = Font(bold=True, name="Arial")
        vc.alignment = Alignment(horizontal="center")
        for cell in [lc, vc]:
            cell.fill  = PatternFill("solid", start_color=bg)
            cell.border = THIN

    wb.save(out)
    print(f"\n[✓] Results exported → {out}")


# ─────────────────────────────────────────────────────────────
# 8. STATISTICAL SUMMARY
# ─────────────────────────────────────────────────────────────

def statistical_summary(df):
    print("\n" + "="*60)
    print("  STATISTICAL RISK ANALYSIS")
    print("="*60)
    feat_stats = df[FEATURE_COLS].describe().round(3)
    print(feat_stats.to_string())

    print("\n  Correlation with Disruption (top 8):")
    corr = df[FEATURE_COLS + ["Disruption_Occurred"]].corr()["Disruption_Occurred"]
    corr = corr.drop("Disruption_Occurred").abs().sort_values(ascending=False).head(8)
    print(corr.to_string())

    print("\n  Risk by Country (avg composite score):")
    country_risk = (df.groupby("Country")["Composite_Risk_Score"]
                      .mean().sort_values(ascending=False).head(8))
    print(country_risk.round(3).to_string())
    print("="*60)


# ─────────────────────────────────────────────────────────────
# 9. MAIN PIPELINE
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n" + "="*60)
    print("  SUPPLY CHAIN RISK PREDICTION SYSTEM")
    print("  Data Science Internship Project")
    print("="*60 + "\n")

    # Step 1: Load
    df, sups = load_and_explore("supply_chain_data.xlsx")

    # Step 2: Feature Engineering
    df = engineer_features(df)

    # Step 3: Statistical Summary
    statistical_summary(df)

    # Step 4: Prepare
    (X_train, X_test, y_train, y_test,
     Xm_train, Xm_test, ym_train, ym_test,
     scaler, le, X_scaled, X_orig) = prepare_data(df)

    # Step 5: Train Binary Models
    results, best_name = train_models(X_train, X_test, y_train, y_test)

    # Step 6: Train Multiclass Model
    xgb_multi = train_multiclass(Xm_train, Xm_test, ym_train, ym_test, le)

    # Step 7: Feature Importance
    importance = feature_importance(results, best_name)

    # Step 8: Score All Records
    scored_df = score_all_suppliers(df, results, best_name, scaler)

    # Step 9: Export
    export_results(scored_df, results, best_name, importance,
                   "supply_chain_risk_results.xlsx")

    print("\n[✓] Pipeline complete!")
    print("    Files generated:")
    print("      supply_chain_data.xlsx         ← Input dataset")
    print("      supply_chain_risk_results.xlsx ← Full results & model outputs\n")
